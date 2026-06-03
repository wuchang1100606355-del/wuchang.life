#!/usr/bin/env python3
from pathlib import Path
import openpyxl
import json

SOURCE = Path("/home/taiji_01/Wuchang_Project/匯出菜單-聊閣社區咖啡重新店-QC_1760535925901.xlsx")
OUT = Path("import/quickclick_xlsx_inspect_report.json")
OUT.parent.mkdir(parents=True, exist_ok=True)

if not SOURCE.exists():
    raise SystemExit(f"CRITICAL: source xlsx not found: {SOURCE}")

wb = openpyxl.load_workbook(SOURCE, data_only=True, read_only=True)

report = {
    "source": str(SOURCE),
    "sheets": []
}

for ws in wb.worksheets:
    rows = []
    non_empty_count = 0

    for r_idx, row in enumerate(ws.iter_rows(values_only=True), 1):
        values = ["" if v is None else str(v).strip() for v in row]
        if any(values):
            non_empty_count += 1
            if len(rows) < 30:
                rows.append({
                    "row": r_idx,
                    "values": values
                })

    report["sheets"].append({
        "sheet": ws.title,
        "max_row": ws.max_row,
        "max_column": ws.max_column,
        "non_empty_rows": non_empty_count,
        "preview": rows
    })

OUT.write_text(json.dumps(report, ensure_ascii=False, indent=2), encoding="utf-8")

print("=== XLSX INSPECT DONE ===")
print("SOURCE=", SOURCE)
print("REPORT=", OUT)
print()
for s in report["sheets"]:
    print("=== SHEET ===", s["sheet"])
    print("max_row:", s["max_row"], "max_column:", s["max_column"], "non_empty_rows:", s["non_empty_rows"])
    for row in s["preview"][:10]:
        print(row["row"], row["values"])
    print()
