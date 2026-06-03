#!/usr/bin/env python3
# -*- coding: utf-8 -*-
import os
import csv
from pathlib import Path
try:
    import openpyxl
except ImportError:
    print("ERROR: 缺少 openpyxl")
    exit(1)

SOURCE_XLSX = Path("/home/taiji_01/Wuchang_Project/匯出菜單-聊閣社區咖啡重新店-QC_1760535925901.xlsx")
IMPORT_DIR = Path("/home/taiji_01/Taiji_Hub/Taiji_Odoo/import")
IMPORT_DIR.mkdir(parents=True, exist_ok=True)

OUT_MENU = IMPORT_DIR / "quickclick_menu_raw.tsv"
OUT_GROUPS = IMPORT_DIR / "quickclick_option_groups.tsv"
OUT_ITEMS = IMPORT_DIR / "quickclick_option_items.tsv"

def clean_data():
    if not SOURCE_XLSX.exists():
        print(f"CRITICAL ERROR: 找不到母檔 {SOURCE_XLSX}")
        exit(1)

    print(f"[*] 載入母檔 (啟用 read_only 極速模式): {SOURCE_XLSX.name}")
    wb = openpyxl.load_workbook(SOURCE_XLSX, data_only=True, read_only=True)
    
    print("[*] 執行業務邏輯：降維(無變體)、中杯計價、耶加雪夫/曼特寧 870元校正")
    menu_data = []
    ws_menu = wb.worksheets[0] 
    
    headers = []
    for row in ws_menu.iter_rows(min_row=1, max_row=1, values_only=True):
        headers = [str(cell).strip() if cell else "" for cell in row]
        break
    
    name_idx = headers.index("主商品名稱") if "主商品名稱" in headers else 0
    cat_idx = headers.index("主商品類別") if "主商品類別" in headers else 1
    price_idx = headers.index("主商品價格") if "主商品價格" in headers else 2
    
    processed_names = set()
    empty_row_count = 0
    
    for row in ws_menu.iter_rows(min_row=2, values_only=True):
        if not row or not row[name_idx]: 
            empty_row_count += 1
            if empty_row_count > 20: break
            continue
            
        empty_row_count = 0
        
        name = str(row[name_idx]).strip()
        category = str(row[cat_idx]).strip() if row[cat_idx] else "未分類"
        price = row[price_idx] or 0
        
        try: price = int(float(price))
        except ValueError: price = 0

        if "簡餐" in category or "定食" in category:
            valid_drinks = {"錫蘭紅茶": 0, "茉香綠茶": 0, "伯爵紅茶": 5}
            matched = False
            for drink, delta in valid_drinks.items():
                if drink in name:
                    name = drink; price = delta; matched = True; break
            if not matched: continue 

        if ("耶加雪夫" in name or "曼特寧" in name) and "磅" in name:
            price = 870

        base_name = name.replace("(大杯)", "").replace("(中杯)", "").replace("(大)", "").replace("(中)", "").strip()
        if base_name in processed_names: continue
        processed_names.add(base_name)
        
        menu_data.append({"主商品名稱": base_name, "主商品類別": category, "基準價格": price})

    wb.close()

    with open(OUT_MENU, 'w', encoding='utf-8', newline='') as f:
        writer = csv.writer(f, delimiter='\t')
        writer.writerow(["主商品名稱", "主商品類別", "基準價格"])
        for d in menu_data: writer.writerow([d["主商品名稱"], d["主商品類別"], d["基準價格"]])
    print(f"[+] {OUT_MENU.name} 匯出完成 (共 {len(menu_data)} 筆無變體商品)")

    print("[*] 提取選項群組與細項...")
    with open(OUT_GROUPS, 'w', encoding='utf-8', newline='') as f:
        f.write("題型選項組合編號\t題型選項組合名稱\n")
        groups = [("O7835309", "尺寸(30)+溫度+甜度"), ("O7835310", "尺寸(35)+溫度+甜度"), ("O7835311", "尺寸(40)+溫度+甜度"), ("O7835312", "尺寸(5)+溫度+甜度"), ("O7835313", "尺寸(10)+溫度+甜度"), ("O7835314", "尺寸(15)+溫度+甜度"), ("O7835315", "貝果口味"), ("O7835316", "尺寸(25)+溫度+甜度"), ("O7835317", "黃金曼特寧+耶加雪夫"), ("O7835318", "曼特寧+藍山"), ("O7835319", "曼巴"), ("O7835320", "招牌咖啡豆"), ("O7835321", "尺寸(40)+糖漿口味+溫度+甜度"), ("O7835325", "尺寸(20)+溫度+甜度"), ("O7835326", "厚片口味"), ("O7835329", "簡餐飲品"), ("O8536132", "特調加購"), ("O8640678", "手沖溫控方式"), ("O8701672", "定食飲料"), ("O8701886", "加購鮮奶咖啡"), ("O8796286", "西西里調整項目")]
        for g in groups: f.write(f"{g[0]}\t{g[1]}\n")

    with open(OUT_ITEMS, 'w', encoding='utf-8', newline='') as f:
        f.write("選項所屬群組\t選項名稱\t加價\n")
        f.write("O7835329\t錫蘭紅茶\t0\n")
        f.write("O7835329\t茉香綠茶\t0\n")
        f.write("O7835329\t伯爵紅茶\t5\n")
    print(f"[+] {OUT_GROUPS.name} 與 {OUT_ITEMS.name} 匯出完成")

if __name__ == "__main__":
    clean_data()
